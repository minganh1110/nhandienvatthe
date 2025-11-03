# -*- coding: utf-8 -*-
import os, time, glob, subprocess, threading, queue, json, sys, random
from pathlib import Path

# ---- GUI / CV / IO
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import cv2
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from nsfw_detector import predict
# ================== DEFAULT CONFIG (override bởi settings.json) ==================
EXCEL_FILE_DEFAULT   = r"C:\toollinkanh\nhandienvatthe\link.xlsx"
DOWNLOAD_DIR_DEFAULT = r"C:\Users\ADMIN\Downloads\fileanhtam"
IMG_DONE_DIR_DEFAULT = r"C:\toollinkanh\nhandienvatthe\imgdone"
RESULT_XLSX_DEFAULT  = str(Path(IMG_DONE_DIR_DEFAULT) / "result_links.xlsx")

CHROME_EXE           = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
USER_DATA_DIR        = r"C:\Users\dell\AppData\Local\Google\Chrome\User Data"
PROFILE_DIR_DEFAULT  = "Profile 47"

CONF_THRES   = 0.5
IDEAL_RATIO  = 0.4
OPEN_LINK_DELAY = 3.2
MAX_WAIT_DL     = 120
QUIET_SECONDS   = 2.0
IMG_EXTS        = (".jpg", ".jpeg", ".png", ".webp")
CLOSE_TAB_EACH_ROUND = True

# ====== EAST (text detector) — load once ======
EAST_MODEL = "frozen_east_text_detection.pb"
if not os.path.exists(EAST_MODEL):
    raise FileNotFoundError("❌ Không tìm thấy file 'frozen_east_text_detection.pb' trong thư mục hiện tại!")
NET_EAST = cv2.dnn.readNet(EAST_MODEL)

# ================== KẾT QUẢ EXCEL: 3 cột ==================
HEADERS = ["stt", "link sản phẩm", "stt ảnh đã chọn"]
SETTINGS_PATH = Path(__file__).with_name("settings.json")

# ================== DETERMINISM ==================
def set_deterministic(seed=1234):
    random.seed(seed)
    np.random.seed(seed)

# ================== UTIL: load/save settings ==================
def load_settings():
    s = {
        "excel_file": EXCEL_FILE_DEFAULT,
        "download_dir": DOWNLOAD_DIR_DEFAULT,
        "img_done_dir": IMG_DONE_DIR_DEFAULT,
        "result_xlsx": RESULT_XLSX_DEFAULT,
        "profile_dir_name": PROFILE_DIR_DEFAULT
    }
    if SETTINGS_PATH.exists():
        try:
            s.update(json.loads(SETTINGS_PATH.read_text(encoding="utf-8")))
        except Exception:
            pass
    if not s.get("result_xlsx"):
        s["result_xlsx"] = str(Path(s["img_done_dir"]) / "result_links.xlsx")
    return s

def save_settings(d):
    keep = {k: d[k] for k in ["excel_file","download_dir","img_done_dir","result_xlsx","profile_dir_name"] if k in d}
    SETTINGS_PATH.write_text(json.dumps(keep, ensure_ascii=False, indent=2), encoding="utf-8")

# ================== FILE / EXCEL helpers ==================
def ensure_dirs(download_dir, img_done_dir):
    os.makedirs(download_dir, exist_ok=True)
    os.makedirs(img_done_dir, exist_ok=True)

def clear_download_dir(download_dir):
    for pattern in [*[f"*{ext}" for ext in IMG_EXTS], "*.crdownload"]:
        for p in glob.glob(os.path.join(download_dir, pattern)):
            try: os.remove(p)
            except: pass

def excel_init(path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    wb.save(path); wb.close()

def excel_ensure(path):
    if not os.path.exists(path):
        excel_init(path)

def excel_next_stt(path):
    excel_ensure(path)
    wb = load_workbook(path); ws = wb.active
    n = ws.max_row
    wb.close()
    return n

def excel_append_row(path, row3):
    excel_ensure(path)
    wb = load_workbook(path); ws = wb.active
    ws.append(row3); wb.save(path); wb.close()

def excel_clear_file(path):
    excel_init(path)

def read_links(excel_path):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(excel_path)
    df = pd.read_excel(excel_path, engine="openpyxl")
    col = "link" if "link" in df.columns else df.columns[0]
    return [str(x).strip() for x in df[col].tolist() if str(x).strip()]

# ================== DOWNLOAD WATCHER ==================
def list_images(dirpath):
    res = []
    for ext in IMG_EXTS:
        res += glob.glob(os.path.join(dirpath, f"*{ext}"))
    return res

def has_crdownload(dirpath):
    return bool(glob.glob(os.path.join(dirpath, "*.crdownload")))

def wait_download_batch(download_dir, timeout=120, quiet_seconds=8.0):
    t0 = time.time()
    last_sig, last_change = None, time.time()
    def snapshot():
        fs = list_images(download_dir)
        sig = tuple(sorted((os.path.basename(p).lower(), os.path.getsize(p)) for p in fs))
        return fs, sig
    while time.time() - t0 < timeout:
        if has_crdownload(download_dir):
            last_sig = None; last_change = time.time()
            time.sleep(0.3); continue
        files, sig = snapshot()
        if sig != last_sig:
            last_sig, last_change = sig, time.time()
        else:
            if time.time() - last_change >= quiet_seconds:
                time.sleep(1.0)
                return list_images(download_dir)
        time.sleep(0.3)
    return list_images(download_dir)

# ================== TEXT RATIO (EAST) ==================
def text_ratio_east(img, conf_threshold=0.5):
    """Tỷ lệ vùng chữ 0..1 (EAST)."""
    h, w = img.shape[:2]
    new_w, new_h = 320, 320
    blob = cv2.dnn.blobFromImage(
        cv2.resize(img, (new_w, new_h)), 1.0, (new_w, new_h),
        (123.68, 116.78, 103.94), swapRB=True, crop=False
    )
    NET_EAST.setInput(blob)
    scores, geometry = NET_EAST.forward(
        ["feature_fusion/Conv_7/Sigmoid", "feature_fusion/concat_3"]
    )
    numRows, numCols = scores.shape[2:4]
    boxes, confs = [], []
    for y in range(numRows):
        s = scores[0, 0, y]
        x0, x1, x2, x3, ang = (geometry[0, i, y] for i in range(5))
        for x in range(numCols):
            if s[x] < conf_threshold: continue
            offX, offY = x*4.0, y*4.0
            c, sn = float(np.cos(ang[x])), float(np.sin(ang[x]))
            hbox = x0[x] + x2[x]; wbox = x1[x] + x3[x]
            endX = int(offX + c*x1[x] + sn*x2[x])
            endY = int(offY - sn*x1[x] + c*x2[x])
            startX = int(endX - wbox)
            startY = int(endY - hbox)
            boxes.append((startX, startY, endX, endY)); confs.append(float(s[x]))
    if not boxes:
        return 0.0
    rects = [cv2.boundingRect(np.array([[x1,y1],[x2,y2]])) for (x1,y1,x2,y2) in boxes]
    idxs  = cv2.dnn.NMSBoxes(rects, confs, conf_threshold, 0.4)
    if len(idxs) == 0:
        return 0.0
    text_area = 0
    for i in idxs.flatten():
        x, y, ww, hh = rects[i]
        text_area += max(0, ww) * max(0, hh)
    return min(1.0, text_area / float(h*w))

EPS = 1e-6
def stable_key(p):
    try:
        return (os.path.basename(p).lower(), os.path.getsize(p))
    except Exception:
        return (os.path.basename(p).lower(), 0)

# ================== MODEL YOLO ==================
from ultralytics import YOLO
import tensorflow as tf
YOLO_MODEL_PATH = "yolov8n.pt"
if not os.path.exists(YOLO_MODEL_PATH):
    raise FileNotFoundError("❌ Không tìm thấy file YOLO model (.pt)!")
MODEL_YOLO = YOLO(YOLO_MODEL_PATH)

# ================== MODEL EAST ==================
EAST_MODEL = "frozen_east_text_detection.pb"
if not os.path.exists(EAST_MODEL):
    raise FileNotFoundError("❌ Không tìm thấy file 'frozen_east_text_detection.pb'!")
net = cv2.dnn.readNet(EAST_MODEL)

# ================== MODEL NSFW ==================
H5_NAME = "nsfw_mobilenet2.224x224.h5"
SAVED_DIR = "nsfw_mobilenet2.224x224_saved"

def ensure_nsfw_model():
    if os.path.isdir(SAVED_DIR):
        return predict.load_model(SAVED_DIR)
    if os.path.isfile(H5_NAME):
        keras_model = tf.keras.models.load_model(H5_NAME)
        keras_model.save(SAVED_DIR)
        return predict.load_model(SAVED_DIR)
    raise FileNotFoundError("Không có NSFW model (.h5 hoặc .saved)")

NSFW_MODEL = ensure_nsfw_model()


# ================== HÀM PHỤ ==================
def skin_ratio(img):
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    lower = np.array([0, 40, 60], dtype=np.uint8)
    upper = np.array([20, 150, 255], dtype=np.uint8)
    mask = cv2.inRange(hsv, lower, upper)
    return np.sum(mask > 0) / mask.size


def is_sensitive_image(path):
    preds = predict.classify(NSFW_MODEL, path)
    result = preds[path]
    img = cv2.imread(path)
    if img is None:
        return False

    skin = skin_ratio(img)
    porn = result.get("porn", 0)
    sexy = result.get("sexy", 0)
    hentai = result.get("hentai", 0)

    if (porn > 0.6 or sexy > 0.6 or hentai > 0.6) and skin > 0.2:
        print(f"🚫 Ảnh nhạy cảm bị loại: {os.path.basename(path)} ({result}, skin={skin:.2f})")
        return True
    if skin > 0.45:
        print(f"🚫 Ảnh có vùng da lớn ({skin*100:.1f}%) => loại")
        return True
    return False


def sharpness_score(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return cv2.Laplacian(gray, cv2.CV_64F).var()


def focus_ratio(img):
    results = MODEL_YOLO(img, verbose=False)
    h, w = img.shape[:2]
    total_area = w * h
    object_area = 0
    for box in results[0].boxes.xyxy.cpu().numpy():
        x1, y1, x2, y2 = box[:4]
        object_area += max(0, (x2 - x1)) * max(0, (y2 - y1))
    return min(object_area / total_area, 1.0)


def text_ratio(img, conf_threshold=0.5):
    h, w = img.shape[:2]
    blob = cv2.dnn.blobFromImage(cv2.resize(img, (320, 320)), 1.0, (320, 320),
                                 (123.68, 116.78, 103.94), swapRB=True, crop=False)
    net.setInput(blob)
    (scores, geometry) = net.forward(
        ["feature_fusion/Conv_7/Sigmoid", "feature_fusion/concat_3"])
    (numRows, numCols) = scores.shape[2:4]

    boxes, confidences = [], []
    for y in range(numRows):
        scoresData = scores[0, 0, y]
        xData = [geometry[0, i, y] for i in range(5)]
        for x in range(numCols):
            if scoresData[x] < conf_threshold:
                continue
            offsetX, offsetY = x * 4.0, y * 4.0
            angle = xData[4][x]
            cos, sin = np.cos(angle), np.sin(angle)
            h_box = xData[0][x] + xData[2][x]
            w_box = xData[1][x] + xData[3][x]
            endX = int(offsetX + (cos * xData[1][x]) + (sin * xData[2][x]))
            endY = int(offsetY - (sin * xData[1][x]) + (cos * xData[2][x]))
            startX = int(endX - w_box)
            startY = int(endY - h_box)
            boxes.append([startX, startY, int(w_box), int(h_box)])
            confidences.append(float(scoresData[x]))

    indices = cv2.dnn.NMSBoxes(boxes, confidences, conf_threshold, 0.4)
    total_area = w * h
    text_area = 0
    if len(indices) > 0:
        for i in indices.flatten():
            (x, y, bw, bh) = boxes[i]
            text_area += max(0, bw) * max(0, bh)
    return text_area / total_area


# ================== CHỌN ẢNH TỐT NHẤT ==================
def score_and_pick(file_list):
    """
    Trả về (best_path, best_score, best_text_ratio)
    """
    best_score = -1e9
    best_path = None
    best_tr = 0.0

    for p in file_list:
        if is_sensitive_image(p):  # loại ảnh nhạy cảm
            continue

        try:
            img = cv2.imread(p)
            if img is None:
                continue

            tr = text_ratio(img)
            sharp = sharpness_score(img)
            focus = focus_ratio(img)

            # scale điểm sắc nét (log để tránh lệch)
            sharp_scaled = np.log1p(max(0.0, sharp))

            score = (
                (1.0 - tr) * 0.6 +      # ít chữ
                focus * 0.35 +          # tập trung vật thể
                (sharp_scaled / (sharp_scaled + 1e-6)) * 0.05  # nét
            )

            print(f"📷 {os.path.basename(p)} => text={tr:.3f}, focus={focus:.3f}, sharp={sharp:.1f}, score={score:.3f}")

            if score > best_score:
                best_score = score
                best_path = p
                best_tr = tr

        except Exception as e:
            print(f"❌ Lỗi xử lý {p}: {e}")

    if best_path is None:
        print("❌ Không tìm thấy ảnh phù hợp.")
        return None, 0.0, 0.0
    return best_path, best_score, best_tr

# ================== OPEN LINK ==================
def open_in_chrome(url, profile_dir_name):
    subprocess.Popen(
        [
            CHROME_EXE,
            f"--user-data-dir={USER_DATA_DIR}",
            f"--profile-directory={profile_dir_name}",
            "--new-tab",
            url,
        ],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    )

# ================== THREAD RUNNER ==================
class RunnerThread(threading.Thread):
    def __init__(self, ui):
        super().__init__(daemon=True)
        self.ui = ui
        self._stop = threading.Event()
    def stop(self): self._stop.set()

    def run(self):
        s = self.ui.settings
        ensure_dirs(s["download_dir"], s["img_done_dir"])
        excel_ensure(s["result_xlsx"])

        try:
            links = read_links(self.ui.var_excel.get())
        except Exception as e:
            self.ui.log(f"❌ Không đọc được Excel: {e}")
            self.ui.after(0, self.ui.on_thread_done); return

        self.ui.log(f"Tổng {len(links)} link.")
        for idx_link, link in enumerate(links, start=1):
            if self._stop.is_set(): break

            clear_download_dir(s["download_dir"])
            self.ui.log(f"\n[{idx_link}] Mở link: {link}")
            open_in_chrome(link, self.ui.var_profile.get())
            time.sleep(OPEN_LINK_DELAY)

            self.ui.log("… đợi ảnh ổn định trong thư mục download…")
            files = wait_download_batch(s["download_dir"], MAX_WAIT_DL, QUIET_SECONDS)

            stt = excel_next_stt(s["result_xlsx"])

            if not files:
                self.ui.log("  ❌ Không thấy ảnh.")
                excel_append_row(s["result_xlsx"], [stt, link, 0])
                continue

            files = sorted(files, key=stable_key)
            best_path, best_idx, t_ratio = score_and_pick(files)

            if not best_path:
                self.ui.log("  ❌ Không chọn được ảnh.")
                excel_append_row(s["result_xlsx"], [stt, link, 0])
            else:
                try:
                    im = Image.open(best_path).convert("RGB")
                    self.ui.show_preview_image(im)
                except Exception as e:
                    self.ui.log(f"  ⚠️ Lỗi đọc ảnh để preview: {e}")

                self.ui.log(f"  ✔ Ảnh #{best_idx} (tỷ lệ chữ ~ {t_ratio:.2f})")
                excel_append_row(s["result_xlsx"], [stt, link, best_idx])

            clear_download_dir(s["download_dir"])

            if CLOSE_TAB_EACH_ROUND:
                self.ui.try_close_tab()

        self.ui.log("\nHoàn tất!")
        self.ui.after(0, self.ui.on_thread_done)

# ================== APP (UI) ==================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        try:
            if sys.platform.startswith("win"):
                import ctypes
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except: pass

        self.title("Shopee Auto Pick ")
        self.geometry("1000x640")
        self.minsize(900, 560)

        self.settings = load_settings()

        root = ttk.Frame(self); root.pack(fill="both", expand=True)
        root.columnconfigure(0, weight=0)
        root.columnconfigure(1, weight=1)
        root.rowconfigure(0, weight=1)

        left = ttk.Frame(root, width=340)
        left.grid(row=0, column=0, sticky="nsw")
        right = ttk.Frame(root)
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)

        ttk.Label(left, text="Cấu hình", font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w", padx=10, pady=(10,6), columnspan=3)

        self.var_excel = tk.StringVar(value=self.settings["excel_file"])
        row1 = ttk.Frame(left); row1.grid(row=1, column=0, sticky="ew", padx=10)
        ttk.Label(row1, text="Excel link:").pack(side="left")
        ttk.Entry(row1, textvariable=self.var_excel, width=28).pack(side="left", padx=6)
        ttk.Button(row1, text="Chọn…", command=self.pick_excel).pack(side="left")

        self.var_profile = tk.StringVar(value=self.settings["profile_dir_name"])
        row2 = ttk.Frame(left); row2.grid(row=2, column=0, sticky="ew", padx=10, pady=(6,0))
        ttk.Label(row2, text="Chrome profile:").pack(side="left")
        ttk.Entry(row2, textvariable=self.var_profile, width=20).pack(side="left", padx=6)
        ttk.Button(row2, text="Lưu", command=self.save_profile).pack(side="left")

        ttk.Label(left, text=f"DL: {self.settings['download_dir']}", foreground="#666").grid(row=3, column=0, sticky="w", padx=10, pady=(4,0))
        ttk.Label(left, text=f"OUT: {self.settings['img_done_dir']} (lưu Excel 3 cột)", foreground="#666").grid(row=4, column=0, sticky="w", padx=10)

        btns = ttk.Frame(left); btns.grid(row=5, column=0, sticky="w", padx=10, pady=(6,4))
        self.btn_start = ttk.Button(btns, text="▶ Start", command=self.start_run); self.btn_start.pack(side="left", padx=3)
        self.btn_stop  = ttk.Button(btns, text="⏹ Stop",  command=self.stop_run, state="disabled"); self.btn_stop.pack(side="left", padx=3)
        ttk.Button(btns, text="🧹 Clear file", command=self.clear_all).pack(side="left", padx=3)
        ttk.Button(btns, text="↺ Reset",      command=self.reset_tool).pack(side="left", padx=3)

        ttk.Label(left, text="Log:").grid(row=6, column=0, sticky="w", padx=10)
        logwrap = ttk.Frame(left); logwrap.grid(row=7, column=0, sticky="nsew", padx=10, pady=(0,10))
        left.rowconfigure(7, weight=1)
        self.txt = tk.Text(
            logwrap, height=26, width=44,
            bg="#1e293b", fg="#e2e8f0",
            insertbackground="#e2e8f0",
            selectbackground="#334155", selectforeground="#f8fafc",
            wrap="word", borderwidth=0, highlightthickness=0
        )
        vsb = ttk.Scrollbar(logwrap, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=vsb.set)
        self.txt.pack(side="left", fill="both", expand=True); vsb.pack(side="right", fill="y")

        ttk.Label(right, text="Preview ảnh chọn", font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w", padx=10, pady=(10,6))
        self.preview_canvas = tk.Canvas(right, bg="#202020", highlightthickness=0)
        self.preview_canvas.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.preview_img = None
        self.preview_imgtk = None
        self.preview_canvas.bind("<Configure>", self._redraw_preview)

        self.queue = queue.Queue()
        self.runner = None
        self.after(80, self.drain_queue)

    def _redraw_preview(self, _evt=None):
        self.preview_canvas.delete("all")
        cw, ch = self.preview_canvas.winfo_width(), self.preview_canvas.winfo_height()
        if not self.preview_img:
            self.preview_canvas.create_text(cw/2, ch/2, text="(Chưa có ảnh chọn)", fill="#888")
            return
        try:
            im = self.preview_img.copy()
            im.thumbnail((max(1, cw-4), max(1, ch-4)), Image.LANCZOS)
            self.preview_imgtk = ImageTk.PhotoImage(im)
            self.preview_canvas.create_image(cw//2, ch//2, image=self.preview_imgtk)
        except Exception as e:
            self.preview_canvas.create_text(10,10, anchor="nw", text=f"Lỗi ảnh: {e}", fill="#f66")

    def show_preview_image(self, pil_image):
        self.preview_img = pil_image.copy()
        self._redraw_preview()

    def log(self, msg): self.queue.put(msg)

    def drain_queue(self):
        try:
            while True:
                msg = self.queue.get_nowait()
                self.txt.insert("end", msg + "\n"); self.txt.see("end")
        except queue.Empty:
            pass
        self.after(80, self.drain_queue)

    def pick_excel(self):
        f = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel", "*.xlsx")])
        if f:
            self.var_excel.set(f)
            self.settings["excel_file"] = f
            save_settings(self.settings)

    def save_profile(self):
        self.settings["profile_dir_name"] = self.var_profile.get().strip() or PROFILE_DIR_DEFAULT
        save_settings(self.settings)
        self.log(f"Đã lưu profile: {self.settings['profile_dir_name']}")

    def clear_all(self):
        if messagebox.askyesno("Xoá", "Xoá toàn bộ dữ liệu trong file kết quả?"):
            excel_clear_file(self.settings["result_xlsx"])
            self.log("Đã tạo lại header (stt, link sản phẩm, stt ảnh đã chọn).")

    def reset_tool(self):
        if self.runner and self.runner.is_alive():
            r = self.runner; r.stop()
            try: r.join(timeout=1.0)
            except: pass
            self.runner = None
        clear_download_dir(self.settings["download_dir"])
        self.txt.delete("1.0", "end")
        self.preview_img = None; self._redraw_preview()
        self.btn_start.config(state="normal"); self.btn_stop.config(state="disabled")
        self.log("Reset xong.")

    def start_run(self):
        if self.runner and self.runner.is_alive(): return
        self.settings["excel_file"]  = self.var_excel.get().strip()
        self.settings["result_xlsx"] = str(Path(self.settings["img_done_dir"]) / "result_links.xlsx")
        save_settings(self.settings)
        self.txt.delete("1.0", "end")
        self.runner = RunnerThread(self); self.runner.start()
        self.btn_start.config(state="disabled"); self.btn_stop.config(state="normal")

    def stop_run(self):
        if self.runner and self.runner.is_alive():
            r = self.runner; r.stop()
            try: r.join(timeout=1.0)
            except: pass
            self.runner = None
        self.btn_start.config(state="normal"); self.btn_stop.config(state="disabled")
        self.log("Đã dừng tool.")

    def on_thread_done(self):
        self.runner = None
        self.btn_start.config(state="normal"); self.btn_stop.config(state="disabled")

    def try_close_tab(self):
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "w")
        except:
            pass

# ================== MAIN ==================
if __name__ == "__main__":
    set_deterministic(1234)
    app = App()
    ensure_dirs(app.settings["download_dir"], app.settings["img_done_dir"])
    app.mainloop()
